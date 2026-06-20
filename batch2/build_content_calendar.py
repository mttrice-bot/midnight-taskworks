import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Content Calendar"
ws['A1'] = "Content Calendar — Monthly Planner"
ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
headers = ["Date", "Platform", "Content Type", "Topic", "Status", "Notes"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=i, value=h)
    c.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    c.fill = openpyxl.styles.PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ws.column_dimensions[chr(64+i)].width = 18
for r in range(4, 34):
    for c in range(1, 7):
        ws.cell(row=r, column=c).border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin'),
            right=openpyxl.styles.Side(style='thin'))
wb.save("content_calendar.xlsx")
