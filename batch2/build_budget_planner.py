import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Budget"
ws['A1'] = "Personal Budget Planner"
ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
for i, h in enumerate(["Category", "Budgeted", "Actual", "Difference"], 1):
    c = ws.cell(row=3, column=i, value=h)
    c.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    c.fill = openpyxl.styles.PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ws.column_dimensions[chr(64+i)].width = 18
cats = ["Housing", "Food", "Transport", "Utilities", "Entertainment", "Savings", "Other"]
for i, cat in enumerate(cats, 4):
    ws.cell(row=i, column=1, value=cat)
    ws.cell(row=i, column=4, value=f"=B{i}-C{i}")
tr = 4 + len(cats)
ws.cell(row=tr, column=1, value="TOTAL").font = openpyxl.styles.Font(bold=True)
for c in [2,3,4]:
    ws.cell(row=tr, column=c, value=f"=SUM({chr(64+c)}4:{chr(64+c)}{tr-1})")
wb.save("budget_planner.xlsx")
