import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Invoice"

style = Font(bold=True, color="FFFFFF", size=11)
fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

ws.merge_cells('A1:F1')
ws['A1'] = "INVOICE"
ws['A1'].font = Font(bold=True, size=20, color="2F5496")
ws['A3'] = "From:"
ws['A4'] = "Your Name / Company"
ws['A5'] = "Your Email"
ws['A6'] = "Your Phone"
ws['C3'] = "Invoice #:"
ws['C4'] = "Date:"
ws['C5'] = "Due Date:"
ws['E3'] = "Bill To:"
ws['E4'] = "Client Name"
ws['E5'] = "Client Email"

headers = ["Item", "Description", "Qty", "Rate", "Total"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=9, column=i, value=h)
    c.font = style; c.fill = fill; c.border = thin

for r in range(10, 14):
    for c in range(1, 6):
        ws.cell(row=r, column=c).border = thin

ws.cell(row=10, column=1, value="Example Item")
ws.cell(row=10, column=2, value="Service description")
ws.cell(row=10, column=3, value=1)
ws.cell(row=10, column=4, value=100)
ws.cell(row=10, column=5, value="=D10*C10")

ws.cell(row=14, column=4, value="Subtotal:").font = Font(bold=True)
ws.cell(row=14, column=5, value="=SUM(E10:E13)").border = thin
ws.cell(row=15, column=4, value="Total:").font = Font(bold=True, size=12)
ws.cell(row=15, column=5, value="=E14").font = Font(bold=True, size=12)
ws.cell(row=15, column=5).border = thin

for col in 'ABCDEF': ws.column_dimensions[col].width = 18
wb.save("/home/team/shared/invoice_template.xlsx")
print("✅ Invoice template built")
