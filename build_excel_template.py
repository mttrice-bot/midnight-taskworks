#!/usr/bin/env python3
"""Generate a Small Business Expense Tracker Excel template using openpyxl"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Dashboard"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Title
ws.merge_cells('A1:F1')
ws['A1'] = "Small Business Expense Tracker"
ws['A1'].font = Font(bold=True, size=16, color="2F5496")

# Monthly Summary
ws['A3'] = "Monthly Summary"
ws['A3'].font = Font(bold=True, size=13)
ws['A4'] = "Month"
ws['B4'] = "Income"
ws['C4'] = "Expenses"
ws['D4'] = "Net"
for row in [4]:
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

months = ["January", "February", "March", "April", "May", "June"]
for i, month in enumerate(months):
    r = 5 + i
    ws.cell(row=r, column=1, value=month).border = thin_border
    ws.cell(row=r, column=2, value=0).border = thin_border
    ws.cell(row=r, column=3, value=0).border = thin_border
    ws.cell(row=r, column=4).border = thin_border
    ws.cell(row=r, column=4, value=f"=B{r}-C{r}")

# Total row
tr = 5 + len(months)
ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
ws.cell(row=tr, column=1).fill = total_fill
ws.cell(row=tr, column=1).border = thin_border
for col in [2, 3, 4]:
    ws.cell(row=tr, column=col).fill = total_fill
    ws.cell(row=tr, column=col).border = thin_border
ws.cell(row=tr, column=2, value=f"=SUM(B5:B{5+len(months)-1})")
ws.cell(row=tr, column=3, value=f"=SUM(C5:C{5+len(months)-1})")
ws.cell(row=tr, column=4, value=f"=B{tr}-C{tr}")

# Column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15

# Sheet 2: Transactions
ws2 = wb.create_sheet("Transactions")
headers = ["Date", "Category", "Description", "Type", "Amount", "Paid Via"]
for col, h in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    ws2.column_dimensions[get_column_letter(col)].width = 18

# Example rows
examples = [
    ("2026-01-05", "Income", "Client payment - Web project", "Income", 2500, "Bank Transfer"),
    ("2026-01-08", "Software", "Monthly SaaS subscription", "Expense", -50, "Credit Card"),
    ("2026-01-10", "Office", "Notebooks and pens", "Expense", -25, "Cash"),
    ("2026-01-15", "Income", "Consulting call", "Income", 500, "PayPal"),
]
for i, (date, cat, desc, typ, amt, via) in enumerate(examples, 2):
    ws2.cell(row=i, column=1, value=date).border = thin_border
    ws2.cell(row=i, column=2, value=cat).border = thin_border
    ws2.cell(row=i, column=3, value=desc).border = thin_border
    ws2.cell(row=i, column=4, value=typ).border = thin_border
    ws2.cell(row=i, column=5, value=amt).border = thin_border
    ws2.cell(row=i, column=6, value=via).border = thin_border

wb.save("/home/team/shared/business_expense_tracker.xlsx")
print("✅ Excel template built: business_expense_tracker.xlsx")
